from io import BytesIO

import pandas as pd
from openpyxl import Workbook, load_workbook

from app.libs.error_exception import ParameterException, ServerError
from app.settings import basedir
from app.utils.file_handler import make_dir


def read_xlsx(file_path_or_stream, sheet_name: str = '', method='path',
              row_start: int = None, row_end: int = None,
              col_start: int = None, col_end: int = None, )->list[list[str,...]]:
    if method == 'path':
        wb = load_workbook(filename=file_path_or_stream, data_only=True)
    elif method == 'stream':
        wb = load_workbook(filename=BytesIO(file_path_or_stream), data_only=True)
    else:
        raise ParameterException()
    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.active
    table_data = []
    rows = ws.rows[row_start:row_end]
    for row in rows:
        row_data = []
        cols = row[col_start:col_end]
        for col in cols:
            if col.value:
                row_data.append(col.value)
            else:
                row_data.append('')
        if row_data:
            table_data.append(row_data)
    return table_data


def generate_xlsx_file(filename, table_sheets, file_dir: str = None):
    wb = Workbook()
    for table_sheet in table_sheets:
        ws = wb.create_sheet(table_sheet['name'])
        table_data = table_sheet['data']
        rows_length = len(table_data)
        for i in range(rows_length):
            col_length = len(table_data[i])
            for j in range(col_length):
                ws.cell(row=i + 1, column=j + 1, value=table_data[i][j])
    wb.remove(wb['Sheet'])
    if file_dir is not None:
        if make_dir(file_dir):
            if not file_dir.endswith('/'):
                file_dir += '/'
            wb.save(file_dir + filename)
        else:
            raise ServerError()
    else:
        wb.save(basedir + '/temp/' + filename)
    return True


def read_xlsx_by_pandas(file_path: str, sheet_name: int or str = 0, header: list or int = 0,
                        names: list = None, index_col: list = None, use_cols: list = None,
                        squeeze: bool = False, converters: dict = None,
                        skip_header_rows: list or int = None, n_rows: int = None, skip_footer_rows: list or int = None
                        )->tuple[list[list[str, ...]], pd.DataFrame]:
    """
    :params header:表示第几行或者前几行都作为列名
    :params names:是否自定义列名
    :params index_col:用于做索引的列，可以是列的名称，也可以是列的索引值；列可以是单个，也可以是多个，即输入一个列表；
    :params use_cols:选择某几列数据
    :params squeeze:当数据仅包含一列，squeeze为True时，返回Series，squeeze为False时，返回DataFrame
    :params converters:强制规定列数据类型，使用一个字典，{"列名":类型}
    :params skip_header_rows:跳过前n行，[a, b, c]:跳过第a+1,b+1,c+1行（索引从0开始）；注意：首行（即列名）也会被跳过
    :params n_rows:需要读取的行数
    :params skip_footer_rows:跳过末尾n行
    """

    df = pd.read_excel(file_path, sheet_name=sheet_name, header=header,
                       index_col=index_col, names=names, usecols=use_cols,
                       squeeze=squeeze, converters=converters,
                       skiprows=skip_header_rows, nrows=n_rows, skipfooter=skip_footer_rows)
    return df.to_numpy().to_list(), df


def generate_xlsx_by_pandas(file_path, df: pd.DataFrame):
    return df.to_excel(file_path, index=False)
